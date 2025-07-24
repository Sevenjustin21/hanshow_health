# modules/utils.py
from datetime import datetime
from openpyxl import load_workbook
import os

class TimeUtils:
    @staticmethod
    def get_time_period():
        now = datetime.now()
        hour, minute = now.hour, now.minute

        if minute <= 4:
            pass
        elif minute >= 55:
            hour = (hour + 1) % 24
        else:
            return None

        if 6 <= hour <= 11 or 16 <= hour <= 23 or 0 <= hour <= 5:
            return "normal"
        if 12 <= hour <= 15:
            return "with_completed"
        return None

    @staticmethod
    def is_already_written(date_key, hour_str, summary_file):
        if not os.path.exists(summary_file):
            return False
        wb = load_workbook(summary_file)
        ws = wb.active
        for col in range(1, ws.max_column + 1, 2):
            if ws.cell(row=1, column=col).value == date_key:
                for r in range(2, ws.max_row + 1):
                    if hour_str in str(ws.cell(row=r, column=col).value):
                        return True
        return False
