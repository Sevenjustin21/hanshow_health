# modules/extractor.py
import os
import re
import pandas as pd
import time
import random
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

class DataExtractor:
    def __init__(self, config, time_utils):
        self.config = config
        self.time_utils = time_utils

    def extract_and_write(self):
        if not os.path.exists(self.config.DATA_FILE):
            print("❌ 未找到数据文件")
            return

        df = pd.read_excel(self.config.DATA_FILE)
        now = datetime.now()
        logic_hour = now.replace(minute=0, second=0, microsecond=0) if now.minute <= 4 else \
                     (now.replace(minute=0, second=0, microsecond=0) + pd.Timedelta(hours=1))

        summary = {
            "Ahold离线数": None,
            "Delhaize-prd离线数": None,
            "Delhaize未对接门店": None,
            "希腊AB门店未对接": None,
            "希腊AB离线数": None,
            "aldi-sued": None,
            "aldi-usa": None,
            "aldi-au": None,
            "ahold-ab": None,
            "更新时间": logic_hour.strftime("%Y-%m-%d %H:00"),
            "更新中数量": {},
            "更新完成数量": {}
        }

        for _, row in df.iterrows():
            brand = str(row["客户号"]).strip().lower()
            brand_std = self.config.BRAND_ALIASES.get(brand)
            if not brand_std: continue

            ap = str(row["AP"]) if not pd.isna(row["AP"]) else ""
            store = row["门店对接"] if not pd.isna(row["门店对接"]) else None
            updating = row["PS更新数量"] if not pd.isna(row["PS更新数量"]) else None
            updated = row["当日更新"] if not pd.isna(row["当日更新"]) else None

            if brand_std == "Ahold":
                m = re.match(r"(\d+)", ap)
                if m: summary["Ahold离线数"] = int(m.group(1))
            elif brand_std == "Delhaize":
                if store is not None:
                    summary["Delhaize未对接门店"] = int(store)
                m = re.match(r"(\d+)", ap)
                if m: summary["Delhaize离线数"] = int(m.group(1))
            elif brand_std == "ahold-ab":
                if store is not None:
                    summary["希腊AB门店未对接"] = int(store)
                m = re.match(r"(\d+)", ap)
                if m: summary["希腊AB离线数"] = int(m.group(1))
                summary["ahold-ab"] = ap
            elif brand_std in ["aldi-sued", "aldi-usa", "aldi-au"]:
                summary[brand_std] = ap
            if brand_std in self.config.BRAND_ALIASES.values():
                if updating is not None:
                    summary["更新中数量"][brand_std] = updating
                if updated is not None:
                    summary["更新完成数量"][brand_std] = updated

        def build_text(include_completed=False):
            line1 = f"Ahold: 基站离线{summary['Ahold离线数']}台"
            line2 = f"Delhaize: {summary['Delhaize未对接门店']}家门店无对接  Delhaize: {summary['Delhaize离线数']}台基站离线  希腊AB：{summary['希腊AB门店未对接']}家门店未对接，{summary['希腊AB离线数']}台基站离线"
            line3 = f"aldi-sued: {summary['aldi-sued']},  aldi-usa: {summary['aldi-usa']} ,  aldi-au: {summary['aldi-au']}"
            updating = [f"{k}: {v}" for k, v in summary["更新中数量"].items()]
            lines = [line1, line2, line3, f"更新中数量: {', '.join(updating)}"]
            if include_completed:
                completed = [f"{k}: {v}" for k, v in summary["更新完成数量"].items()]
                lines.append(f"更新完成数量: {', '.join(completed)}")
            return "\n".join(lines)

        if self.config.DEBUG_MODE:
            print("🛠 DEBUG 模式运行：允许非整点时间写入 / 跳过重复判断")
            period = "with_completed"
        else:
            period = self.time_utils.get_time_period()
            if not period:
                print("⏱ 当前时间不在整点 ±5 分钟范围内，跳过")
                return
            if self.time_utils.is_already_written(now.strftime("%Y-%m-%d"), summary["更新时间"], self.config.SUMMARY_FILE):
                print(f"⏹ 本小时 {summary['更新时间']} 已写入，跳过本轮")
                return

        text = build_text(include_completed=(period == "with_completed"))
        date_key = now.strftime("%Y-%m-%d")
        time_str = summary["更新时间"]

        if os.path.exists(self.config.SUMMARY_FILE):
            wb = load_workbook(self.config.SUMMARY_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active

        date_col = None
        for col in range(1, ws.max_column + 1, 2):
            if ws.cell(row=1, column=col).value == date_key:
                date_col = col
                break

        if date_col is None:
            date_col = ws.max_column + 1 if ws.max_column > 1 else 1
            ws.cell(row=1, column=date_col).value = date_key
            ws.cell(row=1, column=date_col + 1).value = "内容"

        row = 2
        while ws.cell(row=row, column=date_col).value:
            row += 1

        ws.cell(row=row, column=date_col, value=time_str)
        cell = ws.cell(row=row, column=date_col + 1, value=text)
        cell.font = Font(name='微软雅黑', size=10, bold=False, color="FF000000")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrapText=True)

        wb.save(self.config.SUMMARY_FILE)
        print(f"✅ 写入成功 → {self.config.SUMMARY_FILE}：{date_key} - {time_str}")
